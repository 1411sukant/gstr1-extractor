
8 of 29,208
Macro
Inbox

Sukant Sharma
Attachments
Wed 22 Apr, 18:08 (5 days ago)
 

Mail Delivery Subsystem <mailer-daemon@googlemail.com>
Wed 22 Apr, 18:08 (5 days ago)
to me

Error Icon
Address not found
Your message wasn't delivered to 1411491sukant@gamil.com because the address couldn't be found or is unable to receive email.
The response from the remote server was:
511 sorry, no mailbox here by that name (#5.1.1 - chkuser)


Sukant Sharma
Attachments
Sat 25 Apr, 18:07 (2 days ago)
On Wed, 22 Apr 2026 at 18:08, Sukant Sharma <1411491sukant@gmail.com> wrote:

Mail Delivery Subsystem
Sat 25 Apr, 18:07 (2 days ago)
Address not found Your message wasn't delivered to 1411491sukant@gamil.com because the address couldn't be found or is unable to receive email. The response fro

Sukant Sharma <1411491sukant@gmail.com>
Attachments
09:08 (4 hours ago)
to 1411491sukant

 One attachment
  •  Scanned by Gmail

Mail Delivery Subsystem <mailer-daemon@googlemail.com>
09:08 (4 hours ago)
to me

"""
GST Bulk Extractor  ·  GSTR-1 + GSTR-3B
─────────────────────────────────────────
Multi-user ready: each Streamlit browser session is completely isolated.

Output: single Excel file with 4 sheets
  Sheet 1  GSTR-1       Sales, Exports, CDN, Amendments, Tax Liability
  Sheet 2  3.1(d) RCM   Taxable Value + IGST / CGST / SGST
  Sheet 3  4(C) ITC     Net ITC: IGST / CGST / SGST
  Sheet 4  6.1(A)       Tax paid via ITC: Net Payable + IGST / CGST / SGST
"""

import io
import re
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── PAGE CONFIG ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="GST Bulk Extractor", page_icon="🧾", layout="wide")
st.title("🧾 GST Bulk Extractor — GSTR-1 + GSTR-3B")
st.caption(
    "Upload multiple PDFs for each return type. "
    "Multiple team members can use this at the same time — every session is independent."
)

# ── CONSTANTS ─────────────────────────────────────────────────────────────────
MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December",
]
MONTH_FY_ORDER = {m: (i - 3) % 12 for i, m in enumerate(MONTHS)}

# ── SHARED HELPERS ────────────────────────────────────────────────────────────
def pdf_to_text(file) -> str:
    with pdfplumber.open(file) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def fix_broken_numbers(text: str) -> str:
    prev = None
    while prev != text:
        prev = text
        text = re.sub(r'(\d[\d,]*\.\d+)\n(\d+)', r'\1\2', text)
    text = re.sub(r'(\d+)\n(\d{2})\b', r'\1\2', text)
    return text

def find_amounts(text: str, n: int = 1) -> list:
    vals = re.findall(r"-?[\d,]+\.\d{2}", text)
    result = []
    for v in vals:
        result.append(float(v.replace(",", "")))
        if len(result) == n:
            break
    return result

def section_total(text, header_re, stop_re=None, target_word="total", window=1500) -> float:
    m = re.search(header_re, text, re.IGNORECASE | re.DOTALL)
    if not m:
        return 0.0
    start = m.start()
    end = start + window
    if stop_re:
        s = re.search(stop_re, text[start + 10:], re.IGNORECASE)
        if s:
            end = start + 10 + s.start()
    chunk = text[start:end]
    tm = re.search(target_word, chunk, re.IGNORECASE)
    if not tm:
        return 0.0
    vals = find_amounts(chunk[tm.start():], 1)
    return vals[0] if vals else 0.0

def extract_month(text: str) -> str:
    m = re.search(r"(?:Tax\s+[Pp]eriod|Period)\s+([A-Za-z]+)", text)
    if m:
        return m.group(1).capitalize()
    for mo in MONTHS:
        if re.search(mo, text[:600], re.IGNORECASE):
            return mo
    return "Unknown"

def sort_by_month(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_s"] = df["Month"].map(lambda x: MONTH_FY_ORDER.get(x, 99))
    return df.sort_values("_s").drop(columns=["_s"]).reset_index(drop=True)

def row_amounts(text: str, row_re: str, stop_re: str, count: int = 5) -> list:
    m = re.search(row_re, text, re.IGNORECASE)
    if not m:
        return [0.0] * count
    start = m.start()
    stop_m = re.search(stop_re, text[start + 5:], re.IGNORECASE)
    end = start + 5 + (stop_m.start() if stop_m else 500)
    vals = find_amounts(text[start:end], count)
    while len(vals) < count:
        vals.append(0.0)
    return vals

# ── CUSTOM 9A & 6.1(A) HELPERS ────────────────────────────────────────────────
def extract_9A_amendment(text: str) -> float:
    total = 0.0
    sections = re.split(r"\n\s*9A\s*[-–]", text, flags=re.IGNORECASE)
    
    for sec in sections[1:]:
        m = re.search(
            r"Net\s+differential\s+amount.*?(-?[\d,]+\.\d{2})",
            sec, re.IGNORECASE | re.DOTALL
        )
        if m:
            total += float(m.group(1).replace(",", ""))
            
    return total

def extract_6_1A(file):
    """
    Extract 6.1(A) — Other than reverse charge — data.

    CRITICAL FIX: file.seek(0) is mandatory because parse_gstr3b calls
    pdf_to_text(file) first which exhausts the file cursor. Without seek(0),
    pdfplumber reads nothing and all values silently return 0.

    Column layout (standard GSTR-3B table):
      Col 0: Description
      Col 1: Tax payable
      Col 2: Adjustment of negative liability
      Col 3: Net Tax Payable          ← we want this per row
      Col 4: ITC – Integrated tax     ← IGST row: ITC-IGST used
      Col 5: ITC – Central tax        ← CGST row: ITC-CGST used
      Col 6: ITC – State/UT tax       ← SGST row: ITC-SGST used
      Col 7: ITC – Cess
      Col 8: Tax paid in cash
      ...

    We read DOWNWARD through section (A) rows only:
      Integrated tax row → net=col3, itc_igst=col4
      Central tax row    → net=col3, itc_cgst=col5
      State/UT tax row   → net=col3, itc_sgst=col6
    """

    # ── Step 1: Reset file cursor (ESSENTIAL) ────────────────────────────────
    try:
        file.seek(0)
    except Exception:
        pass

    # Each liability row: Net Payable + 3 ITC sub-columns (IGST/CGST/SGST)
    igst_net=igst_itc_igst=igst_itc_cgst=igst_itc_sgst=0.0
    cgst_net=cgst_itc_igst=cgst_itc_cgst=cgst_itc_sgst=0.0
    sgst_net=sgst_itc_igst=sgst_itc_cgst=sgst_itc_sgst=0.0

    def parse_cell(val):
        """Convert a pdfplumber cell to float. Handles \n splits, dashes, blanks."""
        if val is None:
            return 0.0
        v = str(val).replace(",", "").replace("\n", "").replace(" ", "").strip()
        if v in ("-", "–", "—", "NA", ""):
            return 0.0
        try:
            return float(v)
        except ValueError:
            return 0.0

    try:
        with pdfplumber.open(file) as pdf:
            for page in pdf.pages:
                for tbl in (page.extract_tables() or []):
                    if not tbl:
                        continue

                    # ── Identify if this is the 6.1 Payment of Tax table ──────
                    flat = " ".join(
                        str(c).lower().replace("\n", " ")
                        for row in tbl for c in (row or []) if c
                    )
                    if "paid through itc" not in flat and "payment of tax" not in flat:
                        continue

                    # ── Find column indices from header rows ──────────────────
                    # Defaults match standard GSTR-3B layout
                    net_col  = 3
                    col_igst = 4
                    col_cgst = 5
                    col_sgst = 6

                    for row in tbl[:6]:   # scan first 6 rows for headers
                        if not row:
                            continue
                        for ci, cell in enumerate(row):
                            if not cell:
                                continue
                            cell_s = str(cell).lower().replace("\n", " ").strip()
                            if "net tax" in cell_s and "payable" in cell_s:
                                net_col = ci
                            # ITC sub-columns come AFTER net_col
                            if ci <= net_col:
                                continue
                            if "integrated" in cell_s:
                                col_igst = ci
                            elif "central" in cell_s:
                                col_cgst = ci
                            elif "state" in cell_s or "/ut" in cell_s:
                                col_sgst = ci

                    # ── Walk rows strictly inside section (A) ─────────────────
                    in_section_a = False
                    for row in tbl:
                        if not row:
                            continue
                        row_text = " ".join(
                            str(c).lower().replace("\n", " ") for c in row if c
                        )

                        # Enter section (A)
                        if "other than reverse charge" in row_text:
                            in_section_a = True
                            continue

                        # STOP at section (B) — never read its rows
                        if in_section_a and "reverse charge" in row_text and "other than" not in row_text:
                            break

                        if not in_section_a:
                            continue

                        # Scan ALL cells for description (row[0] can be None in merged cells)
                        desc = " ".join(
                            str(c).lower().replace("\n", " ")
                            for c in row if c is not None
                        )

                        is_igst = "integrated" in desc and "tax" in desc
                        is_cgst = "central" in desc and "tax" in desc and "integrated" not in desc
                        is_sgst = ("state" in desc or "state/ut" in desc) and "tax" in desc and "central" not in desc and "integrated" not in desc

                        if not (is_igst or is_cgst or is_sgst):
                            continue

                        # ── Read net payable (col 3) and correct ITC column ───
                        def gc(idx):
                            return parse_cell(row[idx]) if len(row) > idx else 0.0

                        # Every row: read Net Payable + ALL 3 ITC sub-columns
                        # IGST row: 31,345 via IGST ITC + 5,71,405 via CGST ITC = 6,02,750 ✓
                        if is_igst:
                            igst_net      = gc(net_col)   # col 3: Net Tax Payable
                            igst_itc_igst = gc(col_igst)  # col 4: paid via IGST ITC
                            igst_itc_cgst = gc(col_cgst)  # col 5: paid via CGST ITC (cross-use)
                            igst_itc_sgst = gc(col_sgst)  # col 6: paid via SGST ITC
                        elif is_cgst:
                            cgst_net      = gc(net_col)
                            cgst_itc_igst = gc(col_igst)  # col 4
                            cgst_itc_cgst = gc(col_cgst)  # col 5: paid via CGST ITC
                            cgst_itc_sgst = gc(col_sgst)  # col 6
                        elif is_sgst:
                            sgst_net      = gc(net_col)
                            sgst_itc_igst = gc(col_igst)  # col 4
                            sgst_itc_cgst = gc(col_cgst)  # col 5
                            sgst_itc_sgst = gc(col_sgst)  # col 6: paid via SGST ITC

    except Exception:
        pass

    return (
        igst_net, igst_itc_igst, igst_itc_cgst, igst_itc_sgst,
        cgst_net, cgst_itc_igst, cgst_itc_cgst, cgst_itc_sgst,
        sgst_net, sgst_itc_igst, sgst_itc_cgst, sgst_itc_sgst,
    )

# ── GSTR-1 PARSER ─────────────────────────────────────────────────────────────
def parse_gstr1(file) -> dict:
    text = fix_broken_numbers(pdf_to_text(file))
    month = extract_month(text)

    b2b = section_total(text, r"4A\s*[-–]?\s*Taxable\s+outward\s+supplies\s+made\s+to\s+registered", r"4B\s*[-–]?\s*Taxable")
    b2cs = section_total(text, r"7\s*[-–]?\s*Taxable\s+supplies.*?unregistered", r"8\s*[-–]?\s*Nil")

    exp_6a  = section_total(text, r"6A\s*[–-]?\s*Exports?\s*\(",   r"6B\s*[-–]?\s*Supplies")
    sez_6b  = section_total(text, r"6B\s*[-–]?\s*Supplies.*?SEZ",  r"6C\s*[-–]?\s*Deemed")
    dee_6c  = section_total(text, r"6C\s*[-–]?\s*Deemed\s+Exports", r"7\s*[-–]?\s*Taxable")

    cdn_reg   = section_total(text, r"9B\s*[-–]?\s*Credit/Debit\s+Notes?\s*\(Registered\)", r"9B\s*[-–]?\s*Credit/Debit\s+Notes?\s*\(Unregistered\)", target_word=r"Total\s*[-–]?\s*Net\s+off")
    cdn_unreg = section_total(text, r"9B\s*[-–]?\s*Credit/Debit\s+Notes?\s*\(Unregistered\)", r"9C\s*[-–]?\s*Amended", target_word=r"Total\s*[-–]?\s*Net\s+off")

    amendment_9a = extract_9A_amendment(text)

    igst = cgst = sgst = 0.0
    m = re.search(r"Total\s+Liability\s*\(Outward[^)]+\)\s*([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})", text, re.IGNORECASE)
    if m:
        igst, cgst, sgst = float(m.group(2).replace(",","")), float(m.group(3).replace(",","")), float(m.group(4).replace(",",""))
    else:
        m2 = re.search(r"Total\s+Liability", text, re.IGNORECASE)
        if m2:
            v = find_amounts(text[m2.start(): m2.start()+400], 4)
            if len(v) >= 4:
                igst, cgst, sgst = v[1], v[2], v[3]

    return {
        "Month":              month,
        "File":               file.name,
        "Sales B2B (4A)":     b2b,
        "Sales B2CS (7)":     b2cs,
        "Total Sales":        b2b + b2cs,
        "6A Exports":         exp_6a,
        "6B SEZ":             sez_6b,
        "6C Deemed Export":   dee_6c,
        "Total Exports":      exp_6a + sez_6b + dee_6c,
        "Credit/Debit Notes": cdn_reg + cdn_unreg,
        "Amendment 9A":       amendment_9a,
        "IGST Liability":     igst,
        "CGST Liability":     cgst,
        "SGST Liability":     sgst,
    }

# ── GSTR-3B PARSER ────────────────────────────────────────────────────────────
def parse_gstr3b(file) -> dict:
    raw_text = pdf_to_text(file)
    text  = fix_broken_numbers(raw_text)
    month = extract_month(text)

    # 3.1(d)
    rcm = row_amounts(text, r"\(d\)\s+Inward supplies\s*\(liable to reverse charge\)", r"\(e\)\s+Non.GST", count=5)
    # 4(C)
    itc = row_amounts(text, r"C\.\s+Net ITC available\s*\(A[-–]?B\)", r"\(D\)\s+Other Details", count=4)

    (igst_net, igst_itc_igst, igst_itc_cgst, igst_itc_sgst,
     cgst_net, cgst_itc_igst, cgst_itc_cgst, cgst_itc_sgst,
     sgst_net, sgst_itc_igst, sgst_itc_cgst, sgst_itc_sgst) = extract_6_1A(file)

    return {
        "Month":             month,
        "File":              file.name,
        "RCM Taxable":       rcm[0],
        "RCM IGST":          rcm[1],
        "RCM CGST":          rcm[2],
        "RCM SGST":          rcm[3],
        "ITC IGST":          itc[0],
        "ITC CGST":          itc[1],
        "ITC SGST":          itc[2],
        # 6.1(A) — Net Tax Payable + ITC utilised per tax head (diagonal)
        # 6.1(A) — IGST liability row (Integrated tax)
        "IGST Net Payable":       igst_net,
        "IGST paid via IGST ITC": igst_itc_igst,   # col 4
        "IGST paid via CGST ITC": igst_itc_cgst,   # col 5 (cross-utilisation)
        "IGST paid via SGST ITC": igst_itc_sgst,   # col 6
        # 6.1(A) — CGST liability row
        "CGST Net Payable":       cgst_net,
        "CGST ITC Utilised": cgst_itc,   # col 5: ITC-CGST used to pay CGST
        "SGST Net Payable":  sgst_net,
        "SGST ITC Utilised": sgst_itc,   # col 6: ITC-SGST used to pay SGST
    }

# ── EXCEL BUILDER ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TTL_FONT = Font(bold=True, color="1F4E79", size=12)
RUPEE    = '#,##0.00'

def write_table(ws, title: str, df: pd.DataFrame, start_row: int) -> int:
    ws.cell(start_row, 1, title).font = TTL_FONT
    start_row += 1
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(start_row, ci, col)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    start_row += 1
    for _, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            c = ws.cell(start_row, ci, val)
            if isinstance(val, float):
                c.number_format = RUPEE
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
        start_row += 1
    return start_row + 1

def build_excel(gstr1_rows: list, gstr3b_rows: list) -> bytes:
    wb = Workbook()

    # Sheet 1 — GSTR-1
    ws1 = wb.active
    ws1.title = "GSTR-1"
    if gstr1_rows:
        df1 = sort_by_month(pd.DataFrame(gstr1_rows))
        write_table(ws1, "GSTR-1 Summary (Month-wise)", df1, 1)
        for i, col in enumerate(df1.columns, 1):
            ws1.column_dimensions[get_column_letter(i)].width = max(18, len(str(col)) + 2)

    # Sheets 2-4 — GSTR-3B
    ws2 = wb.create_sheet("3.1(d) RCM")
    ws3 = wb.create_sheet("4(C) Net ITC")
    ws4 = wb.create_sheet("6.1 Payment of Tax")

    if gstr3b_rows:
        df3 = sort_by_month(pd.DataFrame(gstr3b_rows))
        write_table(ws2, "3.1(d) – Inward Supplies Liable to RCM",
                    df3[["Month","File","RCM Taxable","RCM IGST","RCM CGST","RCM SGST"]], 1)
        write_table(ws3, "4(C) – Net ITC Available (A – B)",
                    df3[["Month","File","ITC IGST","ITC CGST","ITC SGST"]], 1)
        write_table(ws4, "6.1(A) – Payment of Tax (Other than Reverse Charge)",
                    df3[["Month","File","IGST Net Payable","IGST paid via IGST ITC","IGST paid via CGST ITC","IGST paid via SGST ITC","CGST Net Payable","CGST paid via IGST ITC","CGST paid via CGST ITC","CGST paid via SGST ITC","SGST Net Payable","SGST paid via IGST ITC","SGST paid via CGST ITC","SGST paid via SGST ITC"]], 1)
        for ws in (ws2, ws3, ws4):
            for i in range(1, 8):
                ws.column_dimensions[get_column_letter(i)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ── UI ────────────────────────────────────────────────────────────────────────
col_l, col_r = st.columns(2)

with col_l:
    st.subheader("📄 GSTR-1 PDFs")
    gstr1_files = st.file_uploader(
        "Multiple months supported",
        type="pdf", accept_multiple_files=True, key="up1")

with col_r:
    st.subheader("📋 GSTR-3B PDFs")
    gstr3b_files = st.file_uploader(
        "Multiple months supported",
        type="pdf", accept_multiple_files=True, key="up2")

st.divider()

if st.button("⚡ Extract & Download Excel", type="primary",
             disabled=(not gstr1_files and not gstr3b_files)):

    gstr1_rows, gstr3b_rows, errors = [], [], []

    with st.spinner("Processing PDFs…"):
        for f in (gstr1_files or []):
            try:
                gstr1_rows.append(parse_gstr1(f))
            except Exception as e:
                errors.append(f"GSTR-1 | {f.name}: {e}")
        for f in (gstr3b_files or []):
            try:
                gstr3b_rows.append(parse_gstr3b(f))
            except Exception as e:
                errors.append(f"GSTR-3B | {f.name}: {e}")

    for err in errors:
        st.error(f"❌ {err}")

    if gstr1_rows:
        st.markdown("### GSTR-1 Summary")
        df1 = sort_by_month(pd.DataFrame(gstr1_rows))
        num_cols = [c for c in df1.columns if c not in ("Month","File")]
        st.dataframe(df1.style.format({c: "₹{:,.2f}" for c in num_cols}),
                     use_container_width=True)

    if gstr3b_rows:
        df3 = sort_by_month(pd.DataFrame(gstr3b_rows))
        
        st.markdown("### 3.1(d) — RCM")
        rcm_df = df3[["Month","File","RCM Taxable","RCM IGST","RCM CGST","RCM SGST"]]
        st.dataframe(rcm_df.style.format({c: "₹{:,.2f}" for c in rcm_df.columns if c not in ("Month","File")}),
                     use_container_width=True)

        st.markdown("### 4(C) — Net ITC Available")
        itc_df = df3[["Month","File","ITC IGST","ITC CGST","ITC SGST"]]
        st.dataframe(itc_df.style.format({c: "₹{:,.2f}" for c in itc_df.columns if c not in ("Month","File")}),
                     use_container_width=True)

        st.markdown("### 6.1(A) — Payment of Tax (Other than Reverse Charge)")
        paid_df = df3[["Month","File","IGST Net Payable","IGST paid via IGST ITC","IGST paid via CGST ITC","IGST paid via SGST ITC","CGST Net Payable","CGST paid via IGST ITC","CGST paid via CGST ITC","CGST paid via SGST ITC","SGST Net Payable","SGST paid via IGST ITC","SGST paid via CGST ITC","SGST paid via SGST ITC"]]
        st.dataframe(paid_df.style.format({c: "₹{:,.2f}" for c in paid_df.columns if c not in ("Month","File")}),
                     use_container_width=True)

    if gstr1_rows or gstr3b_rows:
        excel_bytes = build_excel(gstr1_rows, gstr3b_rows)
        st.success(f"✅ {len(gstr1_rows)} GSTR-1 and {len(gstr3b_rows)} GSTR-3B file(s) processed.")
        st.download_button(
            label="📥 Download Combined Excel (4 sheets)",
            data=excel_bytes,
            file_name="GST_Bulk_Extract.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
